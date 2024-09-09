# -*- coding: utf-8 -*-
"""
Created on Wed Jun 29 18:28:08 2022

@author: rionm
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

excelPath = r"D:\programs\dataForLowResTest\20220722\meanLIV_0722.xlsx"
WB = openpyxl.load_workbook(excelPath)
WS = WB["Sheet1"]
WS.cell(row=1,column=1).value = "horizontal resolution sd"
WS.cell(row=1,column=2).value = "ROI1[80:140,400:460]" #high LIV
WS.cell(row=1,column=3).value = "ROI2[200:260,280:340]" #low LIV
WS.cell(row=1,column=4).value = "ROI3[300:360,400:460]" #high LIV
WS.cell(row=1,column=5).value = "depth resolution sd"
WS.cell(row=1,column=6).value = "ROI1[80:140,400:460]"
WS.cell(row=1,column=7).value = "ROI2[200:260,280:340]"
WS.cell(row=1,column=8).value = "ROI3[300:360,400:460]"


FilterRes_d = [0.695829, 0.376251, 0.273756, 0.218716, 0.183351, 0.158372,
                0.139654, 0.12504, 0.113281, 0.103598]
FilterRes_h = [0.189483, 0.207776, 0.230183, 0.258354, 0.295027, 0.345153, 0.418931,
                0.542315, 0.816268, 1.000000]

inputFilePath = r"D:\programs\dataForLowResTest\20220722"
dataName = "064"
resolutionVariation = 10

#---when apply filter for both resolution-------------------------------
# for dataId1 in range (0, resolutionVariation ):
#     for dataId2 in range (0, resolutionVariation ):
#         resolution1 = FilterRes_d[dataId1]
#         resolution2 = FilterRes_h[dataId2]
#         datafield = inputFilePath  + "\\" +dataName + "_z" + str('{:.06f}'.format(resolution1)) + "x" + str('{:.06f}'.format(resolution2)) +"_abs2_LIV.npy"
#         LIV = np.load(datafield)
#         ROI1 = LIV[0,220:280,130:190] #core 60pix*60pix
#         ROI2 = LIV[0,100:160,100:160] #periphery
#         ROI3 = LIV[0,330:390,150:210] #periphery
        
#         M_1 = np.average(ROI1, axis=None)
#         M_2 = np.average(ROI2, axis=None)
#         M_3 = np.average(ROI3, axis=None)
        
#         WS.cell(row=(dataId1*resolutionVariation+dataId2+2),column=1).value = str('{:.06f}'.format(resolution1))
#         WS.cell(row=(dataId1*resolutionVariation+dataId2+2),column=2).value = str('{:.06f}'.format(resolution2))
        
#         WS.cell(row=(dataId1*resolutionVariation+dataId2+2),column=3).value = M_1
#         WS.cell(row=(dataId1*resolutionVariation+dataId2+2),column=4).value = M_2
#         WS.cell(row=(dataId1*resolutionVariation+dataId2+2),column=5).value = M_3
        
#         print('--------------dataID1: '+str(dataId1)+"    dataID2:" +str(dataId2)+' was processed.------------------')
#-----------------------------------------------------------------------
#----when apply filter for only one resolution-------------------------
for dataId1 in range (0, resolutionVariation ):
    resolution = FilterRes_h[dataId1] #horizontal resolution
    datafileId = inputFilePath  + "\\" +dataName + "_z" + "non" + "x" + str('{:.06f}'.format(resolution)) +"_abs2_LIV.npy"
    LIV = np.load(datafileId)
    #spheroid
    # ROI1 = LIV[0,220:280,130:190] #core 60pix*60pix
    # ROI2 = LIV[0,100:160,100:160] #periphery
    # ROI3 = LIV[0,330:390,150:210] #periphery
    #organoid
    ROI1 = LIV[0,80:140,400:460] 
    ROI2 = LIV[0,200:260,280:340] 
    ROI3 = LIV[0,300:360,400:460] 
    M_1 = np.average(ROI1, axis=None)
    M_2 = np.average(ROI2, axis=None)
    M_3 = np.average(ROI3, axis=None)
    
    WS.cell(row=(dataId1+2),column=1).value = str('{:.06f}'.format(resolution))
    WS.cell(row=(dataId1+2),column=2).value = M_1
    WS.cell(row=(dataId1+2),column=3).value = M_2
    WS.cell(row=(dataId1+2),column=4).value = M_3

for dataId2 in range (0, resolutionVariation ):
    resolution = FilterRes_d[dataId2] #depth resolution
    datafileId = inputFilePath  + "\\" +dataName + "_z" + str('{:.06f}'.format(resolution)) + "x" + "non" +"_abs2_LIV.npy"
    LIV = np.load(datafileId)
    #spheroid
    # ROI1 = LIV[0,220:280,130:190] #core 60pix*60pix
    # ROI2 = LIV[0,100:160,100:160] #periphery
    # ROI3 = LIV[0,330:390,150:210] #periphery
        #organoid
    ROI1 = LIV[0,80:140,400:460] 
    ROI2 = LIV[0,200:260,280:340] 
    ROI3 = LIV[0,300:360,400:460]
    M_1 = np.average(ROI1, axis=None)
    M_2 = np.average(ROI2, axis=None)
    M_3 = np.average(ROI3, axis=None)
    
    WS.cell(row=(dataId2+2),column=5).value = str('{:.06f}'.format(resolution))
    WS.cell(row=(dataId2+2),column=6).value = M_1
    WS.cell(row=(dataId2+2),column=7).value = M_2
    WS.cell(row=(dataId2+2),column=8).value = M_3

WB.save(excelPath)